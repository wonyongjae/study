from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)
#         """
#         2021-11-07 20:05:47.251000
#         =SUM(1,2,3)
#         =AVERAGE(1,2,3)
#         10
#         20
#         =SUM(A4:A5)
#         """

wb = load_workbook("sample_formula.xlsx", data_only=True) # data_only를 추가할시 수식이 아닌 데이터를 가져옴
ws = wb.active

# 수식이 아닌 실제 데이터를 가져옴
# evaluate 되지 않은 상태의 데이터는 None이라고 뜸
# excel 파일을 켜서 '저장'을 해줘야 적용
for row in ws.values:
    for cell in row:
        print(cell)