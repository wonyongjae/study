from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 성적표 만들기
# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터
    ws.append([i, randint(0,100), randint(0,100)])


col_B = ws["B"] # 영어 column만 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 같이 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 1번째 row만 가져오기
# for cell in row_title:
#     print(cell.value)


# row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가져오기
                    # 2에서 6까지 포함 // "기존의 마지막 인덱스에 +1 안해줘도됨"
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") # coordinage를 이용하여 아래와 같이 출력
#         """
#         A2 B2 C2 
#         A3 B3 C3
#         A4 B4 C4
#         A5 B5 C5
#         A6 B6 C6
#         A7 B7 C7
#         A8 B8 C8
#         A9 B9 C9
#         A10 B10 C10
#         A11 B11 C11
#         """
#         xy = coordinate_from_string(cell.coordinate) # A/10, AZ/250 = 이런 식으로 끊어줌
#         print(xy, end=" ")
#         """
#         ('A', 2) ('B', 2) ('C', 2)    
#         ('A', 3) ('B', 3) ('C', 3)    
#         ('A', 4) ('B', 4) ('C', 4)    
#         ('A', 5) ('B', 5) ('C', 5)    
#         ('A', 6) ('B', 6) ('C', 6)    
#         ('A', 7) ('B', 7) ('C', 7)    
#         ('A', 8) ('B', 8) ('C', 8)    
#         ('A', 9) ('B', 9) ('C', 9)    
#         ('A', 10) ('B', 10) ('C', 10) 
#         ('A', 11) ('B', 11) ('C', 11)
#         """
#         print(xy[0], end="") # A
#         print(xy[1], end=" ") # 1
#     print()


# 전체 rows
# print(tuple(ws.rows)) # 한 행씩 가져와서 튜플로 만들어줌
# for row in tuple(ws.rows):
#     print(row[1].value)

# 전체 columns
# print(tuple(ws.columns)) # 한 열씩 가져와서 튜플로 만들어줌
# for column in tuple(ws.columns):
#     print(column[1].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[2].value)

# for column in ws.iter_rows(): # 전체 column
#     print(column[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
# 슬라이싱과 비슷 
# 한계를 지정함 그래프 그릴때 lim과 비슷 (시작 끝의 한계를 정함)
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     # print(row[0].value, row[1].value) # 수학, 영어
#     print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

# 행 기준으로 가져오냐 or 열 기준으로 가져오냐에 따라서 데이터를 가져오는 방향이 다름
# 모든 인자값 4개를 다 입력할 필요 없음. 상황에 따라 필요한 인자값만 입력해서 슬라이싱 가능

wb.save("sample.xlsx")