from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment   # 스타일 설정을 위한 import

wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자색상(빨강), 글자 기울기(italic), 글자 굵기(bold)
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 글자색, 폰트이름(name), 취소선(strike)
c1.font = Font(color="0000FF", size=20, underline="single") # 글자색, 글자크기(size), 밑줄(underline)

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해 초록색 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Alignment = 위치 정렬
        # center, left, right, top, bottom

        if cell.column == 1: # A 번호열은 제외(첫번째 칼럼제외)
            continue
        # cell 이 점수형 데이터이고 90 점 보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록으로 설정
            cell.font = Font(color="FF0000") # 폰트 색상 바꿈

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

wb.save("sample_style.xlsx")