from openpyxl import Workbook
wb = Workbook()
# wb.create_sheet
ws = wb.create_sheet() # 새로운 sheet 생성

ws.title = "MySheet" # sheet 이름 변경

ws.sheet_properties.tabColor = "ff66ff" 
# RGB 형태로 값을 넣으면 탭 색상 변경
# https://www.w3schools.com/colors/colors_rgb.asp

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성
# Sheet 생성시 index 값으로 Sheet 생성 위치 설정

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인
# 프린트 결과 : ['Sheet', 'MySheet', 'NewSheet', 'YourSheet']

# Sheet 복사
new_ws["A1"] = "Test" # A1셀에 데이터 "Test" 입력
target = wb.copy_worksheet(new_ws) # 위 내용의 Sheet 복사
target.title = "Copied Sheet" # 복사한 Sheet한 이름
# 프린트 결과 : ['Sheet', 'MySheet', 'NewSheet', 'YourSheet', 'Copied Sheet']

wb.save("sample.xlsx")